import unittest
import pandas as pd
import numpy as np
import tempfile
import os
from openpyxl import load_workbook
from ctl1_subset import create_subsets_and_save_ranges, create_subsets_based_on_ranges  
class TestFiberDiameterProcessing(unittest.TestCase):
    def setUp(self):
       
        self.df = pd.DataFrame({
            'FiberDiameter_CTL1': np.linspace(0, 100, num=120),
            'FiberDiameter_CTL2': np.random.uniform(0, 100, size=120),
            'CTL2_Ax': np.random.uniform(0, 1, size=120),
            'CTL2_My': np.random.uniform(0, 1, size=120),
            'G-Ratio_CTL2': np.random.uniform(0, 1, size=120),
            'FiberDiameter_CTL3': np.random.uniform(0, 100, size=120),
            'CTL3_Ax': np.random.uniform(0, 1, size=120),
            'CTL3_My': np.random.uniform(0, 1, size=120),
            'G-Ratio_CTL3': np.random.uniform(0, 1, size=120),
            'FiberDiameter_CTL4': np.random.uniform(0, 100, size=120),
            'CTL4_Ax': np.random.uniform(0, 1, size=120),
            'CTL4_My': np.random.uniform(0, 1, size=120),
            'G-Ratio_CTL4': np.random.uniform(0, 1, size=120)
        })

    def test_create_subsets_and_save_ranges(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            file_path = tmp_file.name

        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        ranges = create_subsets_and_save_ranges(self.df, 'FiberDiameter_CTL1', writer, 'CTL1')
        writer.close()

        workbook = load_workbook(file_path)
        created_sheets = workbook.sheetnames
        expected_sheets = [f"CTL1_Subset{i+1}" for i in range(6)]

        
        print("Created sheets:", created_sheets)
        print("Expected sheets:", expected_sheets)
        
       
        self.assertTrue(all(sheet in expected_sheets for sheet in created_sheets))

        
        for sheet_name in created_sheets:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            subset_index = int(sheet_name.split('Subset')[-1]) - 1  
            min_diameter, max_diameter = ranges[subset_index]['Min Fiber Diameter'], ranges[subset_index]['Max Fiber Diameter']
            if not data.empty:
                self.assertAlmostEqual(data['FiberDiameter_CTL1'].min(), min_diameter)
                self.assertAlmostEqual(data['FiberDiameter_CTL1'].max(), max_diameter)

        os.remove(file_path)


    def test_create_subsets_based_on_ranges(self):
        
        ranges = [{'Subset': 'CTL1_Subset1', 'Min Fiber Diameter': 0, 'Max Fiber Diameter': 20},
                {'Subset': 'CTL1_Subset2', 'Min Fiber Diameter': 21, 'Max Fiber Diameter': 40},
                {'Subset': 'CTL1_Subset3', 'Min Fiber Diameter': 41, 'Max Fiber Diameter': 60},
                {'Subset': 'CTL1_Subset4', 'Min Fiber Diameter': 61, 'Max Fiber Diameter': 80},
                {'Subset': 'CTL1_Subset5', 'Min Fiber Diameter': 81, 'Max Fiber Diameter': 100},
                {'Subset': 'CTL1_Subset6', 'Min Fiber Diameter': 101, 'Max Fiber Diameter': 120}]

        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            file_path = tmp_file.name

        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        create_subsets_based_on_ranges(self.df, 'FiberDiameter_CTL2', ranges, writer, 'CTL2')
        writer.close()

        workbook = load_workbook(file_path)
        created_sheets = workbook.sheetnames

        
        subset_sheets = [f"CTL2_Subset{i+1}" for i in range(1, 6)] 
        print("Created sheets:", created_sheets) 
        
        self.assertTrue(all(sheet in created_sheets for sheet in subset_sheets))

        os.remove(file_path)

if __name__ == '__main__':
    unittest.main()
