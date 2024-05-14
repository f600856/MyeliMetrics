import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os




def create_subsets_and_save_ranges(df, column, writer, sheet_name_prefix):
    num_subsets = 6
    total_rows = len(df)
    base_size = int(np.ceil(total_rows / num_subsets))
    last_subset_size = total_rows - base_size * (num_subsets - 1)

    subsets_ranges = []

    for i in range(num_subsets):
        start_index = i * base_size if i < num_subsets - 1 else total_rows - last_subset_size
        end_index = start_index + base_size if i < num_subsets - 1 else start_index + last_subset_size

        subset = df.iloc[start_index:end_index]
        min_fiber_size = subset[column].min() if not subset.empty else None
        max_fiber_size = subset[column].max() if not subset.empty else None

        subset_sheet_name = f"{sheet_name_prefix}_Subset{i+1}"
        subset.to_excel(writer, sheet_name=subset_sheet_name, index=False)

        subsets_ranges.append({
            "Subset": subset_sheet_name,
            "Min Fiber Diameter": min_fiber_size,
            "Max Fiber Diameter": max_fiber_size
        })

    
    return subsets_ranges if sheet_name_prefix == 'CTL1' else None



def create_subsets_based_on_ranges(df, column, ranges, writer, sheet_name_prefix):
    for i, range_dict in enumerate(ranges):
        min_range = range_dict['Min Fiber Diameter']
        max_range = range_dict['Max Fiber Diameter']

        subset_df = df[(df[column] >= min_range) & (df[column] <= max_range)]
        
        subset_sheet_name = f"{sheet_name_prefix}_Subset{i+1}"
        # Write data to sheet regardless of whether it's empty or not
        subset_df.to_excel(writer, sheet_name=subset_sheet_name, index=False)

        if subset_df.empty:
            print(f"No data found for {sheet_name_prefix} in the range {min_range} to {max_range}")

    
    try:
        workbook = writer.book
        if 'Sheet1' in workbook.sheetnames:
            sheet1 = workbook['Sheet1']
            if sheet1.max_row == 1 and sheet1.max_column == 1 and sheet1.cell(row=1, column=1).value is None:
                workbook.remove(sheet1)
    except Exception as e:
        print(f"Error checking/removing empty 'Sheet1': {e}")




def save_ranges(ranges, ranges_file_path, sheet_name_prefix):
    mode = 'a' if os.path.exists(ranges_file_path) else 'w'
    with pd.ExcelWriter(ranges_file_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
        ranges_df = pd.DataFrame(ranges)
        ranges_df.to_excel(writer, sheet_name=f'{sheet_name_prefix}_Ranges', index=False)

    
    workbook = load_workbook(ranges_file_path)
    if 'Sheet1' in workbook.sheetnames:
        sheet1 = workbook['Sheet1']
        
        if sheet1.max_row == 1 and sheet1.max_column == 1 and sheet1['A1'].value is None:
            workbook.remove(sheet1)
            workbook.save(ranges_file_path)
            print('Empty Sheet1 has been removed from the ranges file.')

    print(f"Fiber diameter ranges for {sheet_name_prefix} have been saved to {ranges_file_path}.")

